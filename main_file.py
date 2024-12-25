import openpyxl
import os

# Excel file name
excel_file = 'tft_data.xlsx'

# Check if the Excel file exists; if not, create it with headers
if not os.path.exists(excel_file):
    # Create a new workbook and add headers
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "User Data"
    sheet.append(["Placement", "Composition", "Main Unit", "Augment 1", "Augment 2", "Augment 3", "Top 4", "Win"])  # Add headers
    workbook.save(excel_file)

# Function to insert data into the Excel spreadsheet
def insert_data(placement, composition, main_unit, aug1, aug2, aug3, top4, win):
    # Load the existing workbook
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    # Append the new data as a row
    sheet.append([placement, composition, main_unit, aug1, aug2, aug3, top4, win])
    # Save after appending
    workbook.save(excel_file)
    print("Data saved to Excel successfully!")

# Function to delete the last row in the Excel spreadsheet
def delete_latest_row():
    try:
        # Load the existing workbook
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # Get the number of rows in the sheet
        max_row = sheet.max_row

        # Ensure there's at least one data row (not just the headers)
        if max_row > 1:
            # Delete the last row
            sheet.delete_rows(max_row)
            workbook.save(excel_file)
            print("Latest row deleted successfully!")
        else:
            print("No data rows to delete.")
    except FileNotFoundError:
        print(f"File '{excel_file}' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Main loop to get user input
def main():
    print('Enter your game details: ')
    placement = valid_placement()
    composition = input('Composition (trait# + trait): ')
    main_unit = input('Main Unit + Star LVL: ')
    aug1 = input('First Augment: ')
    aug2 = input('Second Augment: ')
    aug3 = input('Third Augment: ')
    
    # Checks for top 4 and first rates
    top4 = 0
    win = 0
    if placement < 5:
        top4 = 1
    if placement == 1:
        win = 1
        
    # Insert data into the Excel spreadsheet
    insert_data(placement, composition, main_unit, aug1, aug2, aug3, top4, win)

# Gets valid placement input    
def valid_placement():
    while True:
        try:
            placement_input = int(input('Placement: '))
            if placement_input <= 0 or placement_input > 8:
                print('Placements cannot be negative, 0, or greater than 8')
            else:
                return placement_input
        except ValueError:
            print('Invalid input, please try again.')
                   
# Run the main function
if __name__ == '__main__':
    while True:
        user_input = input('What would you like to do? (add/delete/exit) ').lower()
        if 'add' in user_input:
            main()
        elif 'delete' in user_input:
            delete_latest_row()
        elif 'exit' in user_input:
            print("Goodbye!")
            break
        else:
            print("Invalid option. Please type 'add', 'delete', or 'exit'.")